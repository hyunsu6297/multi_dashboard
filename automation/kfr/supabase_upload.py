"""Upload daily KFR Excel downloads to Supabase as immutable snapshots."""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import urllib.error
import urllib.parse
import urllib.request
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


SOURCE_FILES = {
    "mezzanine_price": "메자닌 기준가.xlsx",
    "fund_trades": "전체펀드 매매현황.xlsx",
    "fund_holdings": "전체펀드 보유현황.xlsx",
}
HEADER_ROW = 2
DATA_START_ROW = 4
KST = ZoneInfo("Asia/Seoul")


def previous_business_day(today: date | None = None) -> date:
    current = today or datetime.now(KST).date()
    current -= timedelta(days=1)
    while current.weekday() >= 5:
        current -= timedelta(days=1)
    return current


def required_env(name: str) -> str:
    value = os.environ.get(name, "").strip()
    if not value:
        raise RuntimeError(f"Environment variable {name} is required")
    return value.rstrip("/")


def json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def parse_workbook(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []
    sheet_names: list[str] = []
    try:
        for sheet in workbook.worksheets:
            sheet_names.append(sheet.title)
            header_values = [cell.value for cell in sheet[HEADER_ROW]]
            active_indexes = [index for index, value in enumerate(header_values) if value not in (None, "")]
            if not active_indexes:
                continue
            columns = [str(header_values[index]).strip() for index in active_indexes]
            for row_no, values in enumerate(
                sheet.iter_rows(min_row=DATA_START_ROW, values_only=True),
                start=DATA_START_ROW,
            ):
                selected = [values[index] if index < len(values) else None for index in active_indexes]
                if all(value in (None, "") for value in selected):
                    continue
                rows.append(
                    {
                        "sheet_name": sheet.title,
                        "row_no": row_no,
                        "payload": {
                            column: json_value(value) for column, value in zip(columns, selected, strict=True)
                        },
                    }
                )
    finally:
        workbook.close()
    return sheet_names, rows


class SupabaseRest:
    def __init__(self, url: str, service_role_key: str) -> None:
        self.base_url = f"{url.rstrip('/')}/rest/v1"
        self.headers = {
            "apikey": service_role_key,
            "Authorization": f"Bearer {service_role_key}",
            "Content-Type": "application/json",
        }

    def request(
        self,
        method: str,
        path: str,
        body: Any | None = None,
        prefer: str | None = None,
    ) -> Any:
        headers = dict(self.headers)
        if prefer:
            headers["Prefer"] = prefer
        data = None if body is None else json.dumps(body, ensure_ascii=False).encode("utf-8")
        request = urllib.request.Request(
            f"{self.base_url}/{path}", data=data, headers=headers, method=method
        )
        try:
            with urllib.request.urlopen(request, timeout=60) as response:
                payload = response.read()
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"Supabase HTTP {exc.code}: {detail[:1000]}") from exc
        return json.loads(payload) if payload else None


def upload_file(
    client: SupabaseRest,
    source_key: str,
    path: Path,
    business_date: date,
) -> tuple[int, int, bool]:
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    sheets, rows = parse_workbook(path)
    query = urllib.parse.urlencode(
        {
            "source_key": f"eq.{source_key}",
            "business_date": f"eq.{business_date.isoformat()}",
            "sha256": f"eq.{digest}",
            "select": "id,row_count",
        }
    )
    existing = client.request("GET", f"kfr_source_snapshots?{query}")
    if existing:
        return int(existing[0]["id"]), int(existing[0]["row_count"]), False

    snapshot = client.request(
        "POST",
        "kfr_source_snapshots",
        {
            "source_key": source_key,
            "business_date": business_date.isoformat(),
            "file_name": path.name,
            "sha256": digest,
            "sheet_names": sheets,
            "row_count": len(rows),
        },
        prefer="return=representation",
    )[0]
    snapshot_id = int(snapshot["id"])
    try:
        payload = [{"snapshot_id": snapshot_id, **row} for row in rows]
        for start in range(0, len(payload), 500):
            client.request(
                "POST",
                "kfr_source_rows",
                payload[start : start + 500],
                prefer="return=minimal",
            )
    except Exception:
        client.request("DELETE", f"kfr_source_snapshots?id=eq.{snapshot_id}")
        raise
    return snapshot_id, len(rows), True


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", default=".")
    parser.add_argument("--business-date", default=previous_business_day().isoformat())
    args = parser.parse_args()

    input_dir = Path(args.input_dir).expanduser().resolve()
    business_date = date.fromisoformat(args.business_date)
    client = SupabaseRest(
        required_env("SUPABASE_URL"),
        required_env("SUPABASE_SERVICE_ROLE_KEY"),
    )

    for source_key, file_name in SOURCE_FILES.items():
        path = input_dir / file_name
        if not path.is_file():
            raise FileNotFoundError(f"Missing KFR download: {path}")
        snapshot_id, row_count, created = upload_file(client, source_key, path, business_date)
        action = "uploaded" if created else "already exists"
        print(f"{source_key}: snapshot={snapshot_id}, rows={row_count}, {action}")


if __name__ == "__main__":
    main()

