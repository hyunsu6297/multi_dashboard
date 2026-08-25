"""Restore dashboard build inputs from Supabase snapshots."""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
KFR_MODULE_DIR = REPO_ROOT / "automation" / "kfr"
if str(KFR_MODULE_DIR) not in sys.path:
    sys.path.insert(0, str(KFR_MODULE_DIR))
from kfr_api import SOURCE_TO_API_NAME, api_shape_rows  # noqa: E402


MANUAL_TARGETS = {
    ("stock", "fund_info"): ("펀드 정보.xlsx", 4),
    ("stock", "sector"): ("업종.xlsx", 1),
    ("stock", "stock_holding"): ("주식보유현황.xlsx", 1),
    ("bond", "fund_info"): ("펀드정보.xlsx", 1),
    ("bond", "issuer"): ("발행사.xlsx", 1),
}


def required_env(name: str) -> str:
    value = os.environ.get(name, "").strip()
    if not value:
        raise RuntimeError(f"Environment variable {name} is required")
    return value


class SupabaseRest:
    def __init__(self, url: str, secret_key: str) -> None:
        self.base_url = f"{url.rstrip('/')}/rest/v1"
        self.headers = {"apikey": secret_key, "Content-Type": "application/json"}
        if secret_key.count(".") == 2:
            self.headers["Authorization"] = f"Bearer {secret_key}"

    def get_all(self, table: str, params: dict[str, str]) -> list[dict[str, Any]]:
        query = urllib.parse.urlencode(params, safe=".,()")
        loaded: list[dict[str, Any]] = []
        for start in range(0, 1_000_000, 1000):
            headers = {**self.headers, "Range": f"{start}-{start + 999}"}
            request = urllib.request.Request(
                f"{self.base_url}/{table}?{query}", headers=headers, method="GET"
            )
            try:
                with urllib.request.urlopen(request, timeout=60) as response:
                    page = json.loads(response.read())
            except urllib.error.HTTPError as exc:
                detail = exc.read().decode("utf-8", errors="replace")
                raise RuntimeError(f"Supabase HTTP {exc.code}: {detail[:1000]}") from exc
            loaded.extend(page)
            if len(page) < 1000:
                return loaded
        raise RuntimeError(f"Pagination limit exceeded for {table}")


def write_workbook(
    path: Path,
    sheets: dict[str, list[dict[str, Any]]],
    header_row: int,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, rows in sheets.items():
        sheet = workbook.create_sheet(str(sheet_name)[:31] or "Data")
        columns: list[str] = []
        for row in rows:
            for column in row["payload"]:
                if column not in columns:
                    columns.append(column)
        for column_index, column in enumerate(columns, start=1):
            sheet.cell(row=header_row, column=column_index, value=column)
        for output_row, row in enumerate(rows, start=header_row + 1):
            for column_index, column in enumerate(columns, start=1):
                sheet.cell(row=output_row, column=column_index, value=row["payload"].get(column))
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def is_effective_sheet_name(value: object) -> bool:
    return bool(re.fullmatch(r"\d{4}-\d{2}-\d{2}", str(value or "").strip()))


def choose_latest_manual_sheet(sheets: dict[str, list[dict[str, Any]]]) -> str:
    dated = sorted(name for name in sheets if is_effective_sheet_name(name))
    if dated:
        return dated[-1]
    if "Sheet1" in sheets:
        return "Sheet1"
    return sorted(sheets)[0]


def restore_kfr_json(client: SupabaseRest, output_dir: Path) -> None:
    snapshots = client.get_all(
        "kfr_source_snapshots",
        {
            "select": "id,source_key,business_date,downloaded_at,file_name,row_count,source_format",
            "order": "business_date.desc,downloaded_at.desc",
        },
    )
    latest: dict[str, dict[str, Any]] = {}
    for snapshot in snapshots:
        latest.setdefault(snapshot["source_key"], snapshot)

    output_dir.mkdir(parents=True, exist_ok=True)
    manifest: dict[str, Any] = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "datasets": {},
    }
    for source_key, api_name in SOURCE_TO_API_NAME.items():
        snapshot = latest.get(source_key)
        if not snapshot:
            raise RuntimeError(f"No KFR Partner API JSON snapshot found for {source_key}")
        source_snapshots = [snapshot]
        if source_key in {"fund_prices", "fund_holdings", "fund_trades", "mezzanine_price"}:
            latest_date = date.fromisoformat(snapshot["business_date"])
            cutoff = latest_date - timedelta(days=31)
            seen_dates = {snapshot["business_date"]}
            for candidate in snapshots:
                if candidate["source_key"] != source_key or candidate["business_date"] in seen_dates:
                    continue
                business_date = date.fromisoformat(candidate["business_date"])
                if business_date < cutoff:
                    continue
                seen_dates.add(candidate["business_date"])
                source_snapshots.append(candidate)
        entries = []
        for snapshot in reversed(source_snapshots):
            rows = client.get_all(
                "kfr_source_rows",
                {
                    "select": "sheet_name,row_no,payload",
                    "snapshot_id": f"eq.{snapshot['id']}",
                    "order": "sheet_name.asc,row_no.asc",
                },
            )
            content = api_shape_rows(source_key, [row["payload"] for row in rows])
            business_date = str(snapshot["business_date"])
            file_name = f"{api_name}_{business_date}.json"
            (output_dir / file_name).write_text(
                json.dumps({"content": content, "total_elements": len(content)}, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            entries.append({
                "business_date": business_date,
                "file": file_name,
                "row_count": len(content),
                "snapshot_id": int(snapshot["id"]),
                "downloaded_at": str(snapshot["downloaded_at"]),
                "source_format": str(snapshot.get("source_format") or "legacy_excel"),
            })
        manifest["datasets"][source_key] = entries
        print(f"restored {source_key}: snapshots={len(entries)}, rows={sum(x['row_count'] for x in entries)}")
    (output_dir / "index.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def restore_kfr(
    client: SupabaseRest,
    stock_dir: Path | None = None,
    bond_dir: Path | None = None,
    mezzanine_dir: Path | None = None,
) -> None:
    """Backward-compatible KFR restore entrypoint for live publishers."""
    restore_kfr_json(client, REPO_ROOT / "data" / "kfr")


def restore_manual(client: SupabaseRest, stock_dir: Path, bond_dir: Path) -> None:
    for (domain, file_key), (file_name, header_row) in MANUAL_TARGETS.items():
        rows = client.get_all(
            "manual_file_rows",
            {
                "select": "sheet_name,row_no,payload",
                "domain": f"eq.{domain}",
                "file_key": f"eq.{file_key}",
                "order": "sheet_name.asc,row_no.asc",
            },
        )
        if not rows:
            raise RuntimeError(f"No manual rows found for {domain}/{file_key}")
        sheets: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in rows:
            sheets[row["sheet_name"]].append(row)
        if domain == "stock" and file_key == "fund_info":
            sheet_name = choose_latest_manual_sheet(sheets)
            sheets = {"Sheet1": sheets[sheet_name]}
        target_dir = stock_dir if domain == "stock" else bond_dir
        write_workbook(target_dir / file_name, sheets, header_row=header_row)
        print(f"restored {domain}/{file_key}: rows={sum(len(sheet_rows) for sheet_rows in sheets.values())}")


def restore_mezzanine_manual(client: SupabaseRest, mezzanine_dir: Path) -> None:
    mezzanine_dir.mkdir(parents=True, exist_ok=True)
    rows = client.get_all("manual_file_rows", {
        "select": "sheet_name,row_no,payload", "domain": "eq.mezzanine",
        "file_key": "eq.instrument_info", "order": "row_no.asc",
    })
    if not rows:
        raise RuntimeError("No manual rows found for mezzanine/instrument_info")
    write_workbook(mezzanine_dir / "종목정보.xlsx", {"Sheet1": rows}, header_row=1)
    fund_rows = client.get_all("manual_file_rows", {
        "select": "sheet_name,row_no,payload", "domain": "eq.mezzanine",
        "file_key": "eq.fund_info", "order": "row_no.asc",
    })
    if not fund_rows:
        raise RuntimeError("No manual rows found for mezzanine/fund_info")
    write_workbook(mezzanine_dir / "펀드정보.xlsx", {"Sheet1": fund_rows}, header_row=1)

    additions = client.get_all("manual_file_rows", {
        "select": "row_no,payload", "domain": "eq.mezzanine",
        "file_key": "eq.instrument_additions", "order": "row_no.asc",
    })
    addition_payload = []
    for row in additions:
        payload = dict(row["payload"])
        addition_payload.append({
            "addition_id": payload.pop("addition_id", ""),
            "linked_instrument_id": payload.pop("linked_instrument_id", ""),
            "fields": payload,
        })
    (mezzanine_dir / "instrument_additions.json").write_text(
        json.dumps(addition_payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    overrides = client.get_all("manual_file_rows", {
        "select": "row_no,payload", "domain": "eq.mezzanine",
        "file_key": "eq.instrument_overrides", "order": "row_no.asc",
    })
    override_payload = {}
    for row in overrides:
        payload = dict(row["payload"])
        instrument_id = str(payload.pop("instrument_id", ""))
        if instrument_id:
            override_payload[instrument_id] = payload
    (mezzanine_dir / "instrument_overrides.json").write_text(
        json.dumps(override_payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    delta_rows = client.get_all("mezzanine_delta_history", {
        "select": "business_date,security_code,security_name,fund_name,nav,nav_return,underlying_change_rate,daily_delta,is_valid,source",
        "order": "business_date.asc,security_code.asc,fund_name.asc",
    })
    (mezzanine_dir / "delta_history.json").write_text(
        json.dumps(delta_rows, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"restored mezzanine manual data: instruments={len(rows)}, additions={len(addition_payload)}, overrides={len(override_payload)}, delta_history={len(delta_rows)}")


def restore_global_manual(client: SupabaseRest, global_dir: Path) -> None:
    if not global_dir.exists():
        return

    def manual_rows(file_key: str) -> list[dict[str, Any]]:
        return client.get_all(
            "manual_file_rows",
            {
                "select": "sheet_name,row_no,payload",
                "domain": "eq.global",
                "file_key": f"eq.{file_key}",
                "order": "sheet_name.asc,row_no.asc",
            },
        )

    etf_rows = manual_rows("etf_db")
    if etf_rows:
        write_workbook(global_dir / "ETF정보.xlsx", {"Sheet1": etf_rows}, header_row=1)
        print(f"restored global/etf_db: rows={len(etf_rows)}")

    fund_rows = manual_rows("fund_info")
    if fund_rows:
        write_workbook(global_dir / "펀드정보.xlsx", {"Sheet1": fund_rows}, header_row=1)
        print(f"restored global/fund_info: rows={len(fund_rows)}")

    emp_info_rows = manual_rows("emp_info")
    emp_portfolio_rows = manual_rows("emp_portfolios")
    if emp_info_rows or emp_portfolio_rows:
        workbook = Workbook()
        workbook.remove(workbook.active)
        summary = workbook.create_sheet("전체")
        summary.append(["구분", "원금"])
        emp_names: list[str] = []
        for row in emp_info_rows:
            payload = row["payload"]
            name = str(payload.get("name") or payload.get("emp") or "").strip()
            if not name:
                continue
            emp_names.append(name)
            summary.append([name, payload.get("principal", 0)])

        grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in emp_portfolio_rows:
            payload = row["payload"]
            name = str(payload.get("emp") or "").strip()
            if not name:
                continue
            if name not in emp_names:
                emp_names.append(name)
                summary.append([name, 0])
            grouped[name].append(payload)

        headers = ["종목", "시총", "3M Avg Vol.", "보유수량", "종가", "등락율", "목표비중"]
        for name in emp_names:
            sheet = workbook.create_sheet(str(name)[:31])
            sheet.append(headers)
            for payload in grouped.get(name, []):
                sheet.append([
                    payload.get("security", ""),
                    payload.get("marketCap", 0),
                    payload.get("avgTurnover3m", 0),
                    payload.get("quantity", 0),
                    payload.get("price", 0),
                    payload.get("change", 0),
                    payload.get("targetWeight", 0),
                ])
        workbook.save(global_dir / "EMP보유현황.xlsx")
        print(f"restored global EMP data: emps={len(emp_names)}, rows={len(emp_portfolio_rows)}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--stock-dir", default="apps/stock")
    parser.add_argument("--bond-dir", default="apps/bond")
    parser.add_argument("--mezzanine-dir", default="apps/mezzanine")
    parser.add_argument("--global-dir", default="apps/global")
    parser.add_argument("--kfr-json-dir", default="data/kfr")
    args = parser.parse_args()
    client = SupabaseRest(required_env("SUPABASE_URL"), required_env("SUPABASE_SERVICE_ROLE_KEY"))
    restore_kfr_json(client, Path(args.kfr_json_dir))
    restore_manual(client, Path(args.stock_dir), Path(args.bond_dir))
    restore_mezzanine_manual(client, Path(args.mezzanine_dir))
    restore_global_manual(client, Path(args.global_dir))


if __name__ == "__main__":
    main()


